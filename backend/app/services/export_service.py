import os
import tempfile
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import pandas as pd
import json
from pathlib import Path
from io import BytesIO
import zipfile

# Импорты для экспорта
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None
    get_column_letter = None

try:
    import xlwt
except ImportError:
    xlwt = None

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
except ImportError:
    SimpleDocTemplate = None

from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from ..database import SessionLocal
from .. import crud, schemas
from ..utils.logger import get_logger, performance_logger
from ..utils.exceptions import (
    FileProcessingError, 
    DatabaseError, 
    ExternalServiceError,
    handle_file_processing_error,
    handle_database_error
)

logger = get_logger(__name__)

class ExportService:
    """Сервис экспорта данных в различные форматы"""
    
    SUPPORTED_FORMATS = ['csv', 'xlsx', 'xls', 'json', 'pdf']
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
    
    def get_db(self) -> Session:
        """Получает сессию базы данных"""
        return SessionLocal()
    
    async def export_order_data(self, export_request: schemas.ExportRequest, user_id: int) -> schemas.ExportResponse:
        """Экспортирует данные заказа в указанном формате"""
        start_time = datetime.now()
        
        logger.info(
            "Export order data started",
            user_id=user_id,
            order_id=export_request.order_id,
            export_format=export_request.export_format
        )
        
        db = self.get_db()
        try:
            # Проверяем права доступа к заказу
            if export_request.order_id:
                try:
                    order = crud.order_crud.get_order(db, export_request.order_id)
                    if not order:
                        raise FileProcessingError(
                            message="Заказ не найден",
                            processing_stage="access_check"
                        )
                except SQLAlchemyError as e:
                    raise handle_database_error("get_order", "orders", e)
            else:
                order = None
            
            # Проверка прав доступа
            if order and hasattr(order, 'user_id') and order.user_id != user_id:
                try:
                    user = crud.get_user_by_id(db, user_id)
                    if not user or not getattr(user, 'is_admin', False):
                        raise FileProcessingError(
                            message="Нет прав доступа к этому заказу",
                            processing_stage="authorization_check"
                        )
                except SQLAlchemyError as e:
                    raise handle_database_error("get_user", "users", e)
            
            # Получаем данные для экспорта
            export_data = await self._prepare_export_data(order, export_request, db)
            
            # Экспортируем в указанном формате
            order_number = getattr(order, 'order_number', 'unknown') if order else 'unknown'
            
            logger.info(
                "Starting export to format",
                order_number=order_number,
                format=export_request.export_format,
                changes_count=len(export_data.get('changes', []))
            )
            
            file_path = await self._export_by_format(
                export_data, 
                export_request.export_format, 
                order_number
            )
            
            # Получаем информацию о файле
            try:
                file_size = os.path.getsize(file_path)
                filename = os.path.basename(file_path)
            except OSError as e:
                raise handle_file_processing_error(
                    filename=file_path,
                    stage="file_info_retrieval",
                    original_error=e
                )
            
            # Создаем URL для скачивания
            download_url = f"/api/v1/export/download/{filename}"
            
            execution_time = (datetime.now() - start_time).total_seconds()
            performance_logger.log_api_performance(
                endpoint="export_order_data",
                method="POST",
                response_time=execution_time,
                status_code=200
            )
            
            logger.info(
                "Export completed successfully",
                order_number=order_number,
                filename=filename,
                file_size=file_size,
                execution_time=execution_time
            )
            
            return schemas.ExportResponse(
                download_url=download_url,
                filename=filename,
                file_size=file_size,
                expires_at=datetime.now() + timedelta(hours=24)
            )
            
        except FileProcessingError:
            # Перебрасываем FileProcessingError как есть
            raise
        except DatabaseError:
            # Перебрасываем DatabaseError как есть
            raise
        except Exception as e:
            logger.error(
                "Unexpected error in export_order_data",
                user_id=user_id,
                order_id=export_request.order_id,
                error=str(e),
                error_type=type(e).__name__
            )
            raise handle_file_processing_error(
                filename="export_request",
                stage="unknown",
                original_error=e
            )
        finally:
            db.close()
    
    async def _prepare_export_data(self, order, export_request: schemas.ExportRequest, db: Session) -> Dict[str, Any]:
        """Подготавливает данные для экспорта"""
        
        logger.info(
            "Preparing export data",
            order_id=getattr(order, 'id', None),
            include_changes=export_request.include_changes
        )
        
        try:
            # Базовая информация о заказе
            if order:
                export_data = {
                    'order_info': {
                        'order_number': getattr(order, 'order_number', 'unknown'),
                        'filename': getattr(order, 'original_filename', 'unknown'),
                        'format': getattr(order, 'file_format', 'unknown'),
                        'status': getattr(order, 'status', 'unknown'),
                        'created_at': getattr(order, 'created_at', datetime.now()).isoformat(),
                        'total_rows': getattr(order, 'total_rows', 0),
                        'processed_rows': getattr(order, 'processed_rows', 0),
                        'progress': getattr(order, 'progress_percentage', 0)
                    },
                    'changes': [],
                    'summary': {}
                }
            else:
                export_data = {
                    'order_info': {
                        'order_number': 'no_order',
                        'filename': 'no_order',
                        'format': 'unknown',
                        'status': 'no_order',
                        'created_at': datetime.now().isoformat(),
                        'total_rows': 0,
                        'processed_rows': 0,
                        'progress': 0
                    },
                    'changes': [],
                    'summary': {}
                }
            
            # Получаем изменения если требуется
            if export_request.include_changes and order:
                try:
                    changes = crud.order_change_crud.get_changes_by_order(db, order.id)
                    
                    logger.info(f"Retrieved {len(changes)} changes for order", order_id=order.id)
                    
                    # Применяем фильтры если есть
                    if export_request.filters:
                        original_count = len(changes)
                        changes = self._apply_filters(changes, export_request.filters)
                        logger.info(
                            f"Applied filters: {original_count} -> {len(changes)} changes",
                            filters=export_request.filters
                        )
                    
                    # Преобразуем изменения в удобный формат
                    changes_data = []
                    for change in changes:
                        try:
                            changes_data.append({
                                'row_number': getattr(change, 'row_number', 0),
                                'column_name': getattr(change, 'column_name', ''),
                                'old_value': getattr(change, 'old_value', ''),
                                'new_value': getattr(change, 'new_value', ''),
                                'change_type': getattr(change, 'change_type', ''),
                                'created_at': getattr(change, 'created_at', datetime.now()).isoformat()
                            })
                        except Exception as e:
                            logger.warning(
                                f"Failed to process change record",
                                change_id=getattr(change, 'id', 'unknown'),
                                error=str(e)
                            )
                            continue
                    
                    export_data['changes'] = changes_data
                    
                    # Создаем сводку по типам изменений
                    change_summary = {}
                    for change in changes:
                        change_type = getattr(change, 'change_type', 'unknown')
                        if change_type not in change_summary:
                            change_summary[change_type] = 0
                        change_summary[change_type] += 1
                    
                    export_data['summary'] = {
                        'total_changes': len(changes),
                        'by_type': change_summary,
                        'export_date': datetime.now().isoformat(),
                        'export_format': export_request.export_format
                    }
                    
                except SQLAlchemyError as e:
                    logger.error(f"Database error getting changes", error=str(e))
                    raise handle_database_error("get_changes_by_order", "order_changes", e)
            
            logger.info(
                "Export data prepared successfully",
                changes_count=len(export_data['changes']),
                has_summary=bool(export_data['summary'])
            )
            
            return export_data
            
        except Exception as e:
            logger.error(f"Error preparing export data", error=str(e), error_type=type(e).__name__)
            raise handle_file_processing_error(
                filename="export_data",
                stage="data_preparation",
                original_error=e
            )
    
    def _apply_filters(self, changes: List, filters: Dict[str, Any]) -> List:
        """Применяет фильтры к изменениям"""
        filtered_changes = changes
        
        # Фильтр по типу изменения
        if 'change_type' in filters:
            change_types = filters['change_type']
            if isinstance(change_types, str):
                change_types = [change_types]
            filtered_changes = [c for c in filtered_changes if c.change_type in change_types]
        
        # Фильтр по диапазону строк
        if 'row_range' in filters:
            row_range = filters['row_range']
            min_row = row_range.get('min', 0)
            max_row = row_range.get('max', float('inf'))
            filtered_changes = [c for c in filtered_changes if min_row <= c.row_number <= max_row]
        
        # Фильтр по колонкам
        if 'columns' in filters:
            columns = filters['columns']
            if isinstance(columns, str):
                columns = [columns]
            filtered_changes = [c for c in filtered_changes if c.column_name in columns]
        
        # Фильтр по дате
        if 'date_range' in filters:
            date_range = filters['date_range']
            start_date = datetime.fromisoformat(date_range['start']) if 'start' in date_range else datetime.min
            end_date = datetime.fromisoformat(date_range['end']) if 'end' in date_range else datetime.max
            filtered_changes = [c for c in filtered_changes if start_date <= c.created_at <= end_date]
        
        return filtered_changes
    
    async def _export_by_format(self, data: Dict[str, Any], format_type: str, order_number: str) -> str:
        """Экспортирует данные в указанном формате"""
        
        exporters = {
            'csv': self._export_to_csv,
            'xlsx': self._export_to_xlsx,
            'xls': self._export_to_xls,
            'json': self._export_to_json,
            'pdf': self._export_to_pdf
        }
        
        exporter = exporters.get(format_type)
        if not exporter:
            raise ValueError(f"Неподдерживаемый формат экспорта: {format_type}")
        
        return await exporter(data, order_number)
    
    async def _export_to_csv(self, data: Dict[str, Any], order_number: str) -> str:
        """Экспорт в CSV формат"""
        filename = f"export_{order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        file_path = os.path.join(self.temp_dir, filename)
        
        # Создаем DataFrame из изменений
        if data['changes']:
            df = pd.DataFrame(data['changes'])
            
            # Добавляем информацию о заказе в начало
            order_info_df = pd.DataFrame([{
                'row_number': 'ORDER_INFO',
                'column_name': 'order_number',
                'new_value': data['order_info']['order_number'],
                'change_type': 'info'
            }, {
                'row_number': 'ORDER_INFO',
                'column_name': 'filename',
                'new_value': data['order_info']['filename'],
                'change_type': 'info'
            }, {
                'row_number': 'ORDER_INFO',
                'column_name': 'total_rows',
                'new_value': str(data['order_info']['total_rows']),
                'change_type': 'info'
            }])
            
            # Объединяем данные
            final_df = pd.concat([order_info_df, df], ignore_index=True)
        else:
            # Если нет изменений, экспортируем только информацию о заказе
            final_df = pd.DataFrame([data['order_info']])
        
        # Сохраняем в CSV
        final_df.to_csv(file_path, index=False, encoding='utf-8-sig')
        
        return file_path
    
    async def _export_to_xlsx(self, data: Dict[str, Any], order_number: str) -> str:
        """Экспорт в XLSX формат"""
        if not openpyxl:
            raise ImportError("openpyxl не установлен")
        
        filename = f"export_{order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(self.temp_dir, filename)
        
        # Создаем рабочую книгу
        wb = Workbook()
        
        # Лист с информацией о заказе
        ws_info = wb.active
        if ws_info is not None:
            ws_info.title = "Order Info"
        
        # Заголовки и стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Информация о заказе
        order_info = data['order_info']
        info_data = [
            ['Параметр', 'Значение'],
            ['Номер заказа', order_info['order_number']],
            ['Имя файла', order_info['filename']],
            ['Формат', order_info['format']],
            ['Статус', order_info['status']],
            ['Дата создания', order_info['created_at']],
            ['Всего строк', order_info['total_rows']],
            ['Обработано строк', order_info['processed_rows']],
            ['Прогресс (%)', order_info['progress']]
        ]
        
        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                if ws_info is not None:
                    cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                    if cell is not None and row_idx == 1:  # Заголовок
                        cell.font = header_font
                        cell.fill = header_fill
        
        # Лист с изменениями
        if data['changes']:
            ws_changes = wb.create_sheet(title="Changes")
            
            # Заголовки для изменений
            headers = ['Строка', 'Колонка', 'Старое значение', 'Новое значение', 'Тип изменения', 'Дата']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_changes.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Данные изменений
            for row_idx, change in enumerate(data['changes'], 2):
                ws_changes.cell(row=row_idx, column=1, value=change['row_number'])
                ws_changes.cell(row=row_idx, column=2, value=change['column_name'])
                ws_changes.cell(row=row_idx, column=3, value=change['old_value'])
                ws_changes.cell(row=row_idx, column=4, value=change['new_value'])
                ws_changes.cell(row=row_idx, column=5, value=change['change_type'])
                ws_changes.cell(row=row_idx, column=6, value=change['created_at'])
        
        # Лист со сводкой
        if data['summary']:
            ws_summary = wb.create_sheet(title="Summary")
            
            summary_data = [
                ['Параметр', 'Значение'],
                ['Всего изменений', data['summary']['total_changes']],
                ['Дата экспорта', data['summary']['export_date']],
                ['Формат экспорта', data['summary']['export_format']]
            ]
            
            # Добавляем статистику по типам
            for change_type, count in data['summary']['by_type'].items():
                summary_data.append([f'Изменения типа "{change_type}"', count])
            
            for row_idx, row_data in enumerate(summary_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:  # Заголовок
                        cell.font = header_font
                        cell.fill = header_fill
        
        # Автоподбор ширины колонок
        for ws in wb.worksheets:
            if ws is not None:
                try:
                    # Устанавливаем фиксированную ширину для всех колонок
                    for col_num in range(1, 10):  # Первые 10 колонок
                        if get_column_letter:
                            column_letter = get_column_letter(col_num)
                            ws.column_dimensions[column_letter].width = 20
                except:
                    # Если автоподбор не работает, пропускаем
                    pass
        
        # Сохраняем файл
        wb.save(file_path)
        
        return file_path
    
    async def _export_to_xls(self, data: Dict[str, Any], order_number: str) -> str:
        """Экспорт в XLS формат"""
        if not xlwt:
            raise ImportError("xlwt не установлен")
        
        filename = f"export_{order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xls"
        file_path = os.path.join(self.temp_dir, filename)
        
        # Создаем рабочую книгу
        wb = xlwt.Workbook()
        
        # Стили
        header_style = xlwt.easyxf('font: bold on; pattern: pattern solid, fore_colour blue;')
        
        # Лист с информацией о заказе
        ws_info = wb.add_sheet('Order Info')
        
        order_info = data['order_info']
        info_data = [
            ['Параметр', 'Значение'],
            ['Номер заказа', order_info['order_number']],
            ['Имя файла', order_info['filename']],
            ['Формат', order_info['format']],
            ['Статус', order_info['status']],
            ['Дата создания', order_info['created_at']],
            ['Всего строк', order_info['total_rows']],
            ['Обработано строк', order_info['processed_rows']],
            ['Прогресс (%)', order_info['progress']]
        ]
        
        for row_idx, row_data in enumerate(info_data):
            for col_idx, value in enumerate(row_data):
                style = header_style if row_idx == 0 else xlwt.easyxf()
                ws_info.write(row_idx, col_idx, value, style)
        
        # Лист с изменениями
        if data['changes']:
            ws_changes = wb.add_sheet('Changes')
            
            headers = ['Строка', 'Колонка', 'Старое значение', 'Новое значение', 'Тип изменения', 'Дата']
            for col_idx, header in enumerate(headers):
                ws_changes.write(0, col_idx, header, header_style)
            
            for row_idx, change in enumerate(data['changes'], 1):
                ws_changes.write(row_idx, 0, change['row_number'])
                ws_changes.write(row_idx, 1, change['column_name'])
                ws_changes.write(row_idx, 2, change['old_value'])
                ws_changes.write(row_idx, 3, change['new_value'])
                ws_changes.write(row_idx, 4, change['change_type'])
                ws_changes.write(row_idx, 5, change['created_at'])
        
        # Сохраняем файл
        wb.save(file_path)
        
        return file_path
    
    async def _export_to_json(self, data: Dict[str, Any], order_number: str) -> str:
        """Экспорт в JSON формат"""
        filename = f"export_{order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        file_path = os.path.join(self.temp_dir, filename)
        
        # Подготавливаем данные для JSON
        export_json = {
            'export_info': {
                'order_number': order_number,
                'export_date': datetime.now().isoformat(),
                'format': 'json'
            },
            'order': data['order_info'],
            'changes': data['changes'],
            'summary': data['summary']
        }
        
        # Сохраняем в JSON файл
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_json, f, ensure_ascii=False, indent=2)
        
        return file_path
    
    async def _export_to_pdf(self, data: Dict[str, Any], order_number: str) -> str:
        """Экспорт в PDF формат"""
        if not SimpleDocTemplate:
            raise ImportError("reportlab не установлен")
        
        filename = f"export_{order_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        file_path = os.path.join(self.temp_dir, filename)
        
        # Создаем PDF документ
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        
        # Стили
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Центрирование
        )
        
        # Заголовок
        story.append(Paragraph(f"Отчет по заказу {order_number}", title_style))
        story.append(Spacer(1, 20))
        
        # Информация о заказе
        story.append(Paragraph("Информация о заказе", styles['Heading2']))
        
        order_info = data['order_info']
        order_table_data = [
            ['Параметр', 'Значение'],
            ['Номер заказа', order_info['order_number']],
            ['Имя файла', order_info['filename']],
            ['Формат', order_info['format']],
            ['Статус', order_info['status']],
            ['Дата создания', order_info['created_at']],
            ['Всего строк', str(order_info['total_rows'])],
            ['Обработано строк', str(order_info['processed_rows'])],
            ['Прогресс (%)', f"{order_info['progress']:.1f}%"]
        ]
        
        order_table = Table(order_table_data)
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(order_table)
        story.append(Spacer(1, 20))
        
        # Сводка по изменениям
        if data['summary']:
            story.append(Paragraph("Сводка по изменениям", styles['Heading2']))
            
            summary_table_data = [
                ['Параметр', 'Значение'],
                ['Всего изменений', str(data['summary']['total_changes'])],
                ['Дата экспорта', data['summary']['export_date']]
            ]
            
            # Добавляем статистику по типам
            for change_type, count in data['summary']['by_type'].items():
                summary_table_data.append([f'Изменения типа "{change_type}"', str(count)])
            
            summary_table = Table(summary_table_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
        
        # Таблица изменений (первые 100 записей)
        if data['changes']:
            story.append(Paragraph("Изменения (первые 100 записей)", styles['Heading2']))
            
            changes_table_data = [['Строка', 'Колонка', 'Старое значение', 'Новое значение', 'Тип']]
            
            # Ограничиваем количество записей для PDF
            limited_changes = data['changes'][:100]
            
            for change in limited_changes:
                # Обрезаем длинные значения
                old_val = str(change['old_value'])[:30] + '...' if change['old_value'] and len(str(change['old_value'])) > 30 else str(change['old_value'] or '')
                new_val = str(change['new_value'])[:30] + '...' if change['new_value'] and len(str(change['new_value'])) > 30 else str(change['new_value'] or '')
                
                changes_table_data.append([
                    str(change['row_number']),
                    str(change['column_name'])[:20] + '...' if len(str(change['column_name'])) > 20 else str(change['column_name']),
                    old_val,
                    new_val,
                    str(change['change_type'])
                ])
            
            changes_table = Table(changes_table_data)
            changes_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(changes_table)
            
            if len(data['changes']) > 100:
                story.append(Spacer(1, 10))
                story.append(Paragraph(f"Показано 100 из {len(data['changes'])} изменений. Для полного списка используйте другие форматы экспорта.", styles['Normal']))
        
        # Генерируем PDF
        doc.build(story)
        
        return file_path

    async def export_analytics_report(self, db: Session, filters: schemas.OrderFilters, format_type: str) -> bytes:
        """Экспорт аналитического отчета"""
        try:
            # Получаем аналитические данные
            from ..crud import get_analytics_data
            analytics_data = get_analytics_data(db, filters.date_from, filters.date_to, "day")
            
            # Экспортируем в указанном формате
            if format_type == 'csv':
                return await self._export_analytics_to_csv(analytics_data)
            elif format_type == 'xlsx':
                return await self._export_analytics_to_xlsx(analytics_data)
            else:
                raise ValueError(f"Неподдерживаемый формат: {format_type}")
                
        except Exception as e:
            logger.error(f"Ошибка экспорта аналитики: {str(e)}")
            raise
    
    async def _export_analytics_to_csv(self, analytics_data: Dict[str, Any]) -> bytes:
        """Экспорт аналитики в CSV"""
        import io
        
        output = io.StringIO()
        
        # Записываем сводку
        output.write("Сводка\n")
        output.write(f"Всего заказов,{analytics_data['summary']['total_orders']}\n")
        output.write(f"Общая выручка,{analytics_data['summary']['total_revenue']}\n")
        output.write(f"Средний чек,{analytics_data['summary']['avg_order_value']}\n")
        output.write("\n")
        
        # Записываем данные по типам оплаты
        output.write("Типы оплаты\n")
        output.write("Тип,Количество,Сумма\n")
        for payment in analytics_data['payment_types']:
            output.write(f"{payment['type']},{payment['count']},{payment['total']}\n")
        output.write("\n")
        
        # Записываем топ автоматы
        output.write("Топ автоматы\n")
        output.write("Код автомата,Количество,Сумма\n")
        for machine in analytics_data['top_machines']:
            output.write(f"{machine['machine_code']},{machine['count']},{machine['total']}\n")
        
        return output.getvalue().encode('utf-8')
    
    async def _export_analytics_to_xlsx(self, analytics_data: Dict[str, Any]) -> bytes:
        """Экспорт аналитики в XLSX"""
        if not openpyxl:
            raise ImportError("openpyxl не установлен")
        
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Заголовки и стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Сводка
        if ws is not None:
            ws.title = "Analytics"
            summary_cell = ws.cell(row=1, column=1, value="Сводка")
            if summary_cell is not None:
                summary_cell.font = header_font
                summary_cell.fill = header_fill
            
            cell = ws.cell(row=2, column=1, value="Всего заказов")
            cell = ws.cell(row=2, column=2, value=analytics_data['summary']['total_orders'])
            cell = ws.cell(row=3, column=1, value="Общая выручка")
            cell = ws.cell(row=3, column=2, value=analytics_data['summary']['total_revenue'])
            cell = ws.cell(row=4, column=1, value="Средний чек")
            cell = ws.cell(row=4, column=2, value=analytics_data['summary']['avg_order_value'])
            
            # Типы оплаты
            start_row = 6
            payment_header_cell = ws.cell(row=start_row, column=1, value="Типы оплаты")
            if payment_header_cell is not None:
                payment_header_cell.font = header_font
                payment_header_cell.fill = header_fill
            
            cell = ws.cell(row=start_row + 1, column=1, value="Тип")
            cell = ws.cell(row=start_row + 1, column=2, value="Количество")
            cell = ws.cell(row=start_row + 1, column=3, value="Сумма")
            
            for i, payment in enumerate(analytics_data['payment_types']):
                row = start_row + 2 + i
                cell = ws.cell(row=row, column=1, value=payment['type'])
                cell = ws.cell(row=row, column=2, value=payment['count'])
                cell = ws.cell(row=row, column=3, value=payment['total'])
        
        # Сохраняем в байты
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# Глобальный экземпляр сервиса экспорта
export_service = ExportService()

# Функция для использования в других модулях
async def export_order_data_async(export_request: schemas.ExportRequest, user_id: int) -> schemas.ExportResponse:
    """Асинхронный экспорт данных заказа"""
    return await export_service.export_order_data(export_request, user_id)
